# Dibuat oleh Faqisna Putra Mardhatillah
# Sebagai tugas kuliah pada Mata Kuliah Praktikum Algoritma dan Pemrograman II

import face_recognition
import cv2
import numpy as np
import os
import tkinter as tk
from PIL import Image, ImageTk
import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime

# Inisialisasi Beberapa Variabel
directoryFaces = "Faces"
LokasiFileNamaTerdeteksi = "Absensi.txt"
NamaSaatkosong = ""
List_Nama_Terdeteksi = []
Deduplicate_List_Nama_Terdeteksi = []
face_encoding = []
Size = 0
Mulai_Pengenalan = False
EncodePertama = True

face_locations = []
face_encodings = []
face_names = []
process_this_frame = True

# Load Nama dari File
File_Wajah = os.listdir(directoryFaces)
WordRemove = ".jpg"
List_Nama = [s.replace(WordRemove, "") for s in File_Wajah]
Size = len(List_Nama)
print(List_Nama)

# Aktifkan Pengenalan
def AktifPengenalan():
    global Mulai_Pengenalan
    Mulai_Pengenalan = True
# Nonaktifkan Pengenalan
def NonaktifPengenalan():
    global Mulai_Pengenalan
    Mulai_Pengenalan = False

# Load/Reload Encoding Data Wajah
def LoadEncodingDataWajah():
    face_encoding.clear()
    for i in range(Size):
        LoadedImage = face_recognition.load_image_file("Faces/" + File_Wajah[i])
        face_encoding.append(face_recognition.face_encodings(LoadedImage)[0])

#Aktifkan-Non Tombol Hapus
def AktifkanTombolHapus(event):
    AksiHapusNamaDariList_Grp["state"] = "normal"
def NonaktifTombolHapus():
    AksiHapusNamaDariList_Grp["state"] = "disable"

# Load/Reload Nama dari Nama File
def LoadNama():
    List_Nama.clear()
    File_Wajah = os.listdir(directoryFaces)
    WordRemove = ".jpg"
    List_Nama = [s.replace(WordRemove, "") for s in File_Wajah]
    Size = len(List_Nama)
    print(List_Nama)
    
# Hapus Nama Dari List
def HapusNamaDariList():
    PilihNama = listNamaTerdeteksi_Tk.curselection()
    for i in PilihNama:
        List_Nama_Terdeteksi.remove(str(listNamaTerdeteksi_Tk.get(i)))
    print(List_Nama_Terdeteksi)
    listNamaTerdeteksi_Tk.delete(0, tk.END)
    for i in List_Nama_Terdeteksi:
        listNamaTerdeteksi_Tk.insert("end", i)

# Capture Photo dan Menyimpan
def CaptureDanSimpan(Nama):
    _, frame = video_capture.read()
    NamaGambar = "Faces/" + Nama + ".jpg"
    cv2.imwrite(NamaGambar, frame)
# Tambah Data
def TambahData():
    windowTambahData = tk.Toplevel()
    windowTambahData.title("Tambah Data")
    tk.Label(windowTambahData, text="Masukkan Nama: ").grid(row=0, column=0)
    FormNama = tk.Entry(windowTambahData, width=30)
    FormNama.grid(row=1, column=0)

    def ThisCommand(Nama):
        CaptureDanSimpan(Nama)
        LoadEncodingDataWajah()
        windowTambahData.destroy()

    tk.Button(windowTambahData, text="Tambah", command= lambda: ThisCommand(FormNama.get())).grid(row=2, column=0)
    windowTambahData.mainloop()

# Menyimpan Teks ke File
def TulisNamaKeFile(Teks, LokasiFile):
    BuffSTR = ""
    if os.path.exists(LokasiFile):
        with open(LokasiFile, "r") as file:
            lines = file.readlines()
            lines = [line.strip() for line in lines]
            for i in range(len(lines)):
                BuffSTR = BuffSTR + lines[i] + '\n'

        with open(LokasiFile, "w") as file:
            file.write(BuffSTR + Teks + "\n")
    else:
        with open(LokasiFile, "w") as file:
            file.write(Teks + "\n")

# Menyimpan File Excel
def SimpanExcel(Header, List, ListHadir, filename):
    List.sort()
    ListHadir.sort()
    Panjang = len(List)
    now = datetime.now()
    currentTime = now.strftime("%d-%m-%Y")
    Exc = openpyxl.Workbook()

    if os.path.exists(filename):
        Exc = openpyxl.load_workbook(filename=filename)
    else:
        Exc.save(filename=filename)
        del Exc["Sheet"]

    Wr = Exc.create_sheet(title=str(currentTime))
    Wr.append(Header)
    BlackFill = PatternFill(start_color='000000',
                   end_color='000000',
                   fill_type='solid')
    RedFill = PatternFill(start_color='FF0000',
                   end_color='FF0000',
                   fill_type='solid')
    GreenFill = PatternFill(start_color='00FF00',
                   end_color='00FF00',
                   fill_type='solid')

    Wr["A1"].font = Font(size=12, bold=True, color= "FFFFFF")
    Wr["B1"].font = Font(size=12, bold=True, color= "FFFFFF")
    Wr["C1"].font = Font(size=12, bold=True, color= "FFFFFF")
    Wr['A1'].fill = BlackFill
    Wr['B1'].fill = BlackFill
    Wr['C1'].fill = BlackFill
 
    for i in range(Panjang):
        Wr['A' + str(i + 2)] = str(i + 1)

    for i in range(Panjang):
        Wr['B' + str(i + 2)] = List[i]

    print(List)
    print(ListHadir)

    for i in range(Panjang):
        if List[i] not in ListHadir:
            Wr['C' + str(i + 2)].fill = RedFill
        else:
            Wr['C' + str(i + 2)].fill = GreenFill
    Exc.save(filename)

# Simpan Ke Excel
def SimpanAbsensi():
    SimpanExcel(["No", "Nama", "Hadir/Tidak"], List_Nama, List_Nama_Terdeteksi, "Excel/Excel.xlsx")

# Setup Window
window = tk.Tk()
window.wm_title("Absensi Face Recognition")
window.config(background="#010101")

# Graphic Window
imageFrame = tk.Frame(window, width=1200, height=500)
imageFrame.grid(row=0, column=0, padx=10, pady=2)
#Entry
##Entry Aktif-Nonaktif Tombol Pengenalan
TombolPengenalanDanStop_Grp = tk.Frame(window)
TombolPengenalanDanStop_Grp.grid(row=1, column=0)
##Entry List Nama yang Sudah Diabsensi
ListNamaAbsensi_Grp = tk.Frame(window)
ListNamaAbsensi_Grp.grid(row=0, column=1, sticky='n')
##Entry Aksi Pilih Nama
AksiPilih_Grp = tk.Frame(window)
AksiPilih_Grp.grid(row=0, column=2, sticky='n')

##Image Show
lmain = tk.Label(imageFrame)
lmain.grid(row=0, column=0)
##Label Nama yang Terdeteksi
LabelSudahAbsen_Grp = tk.Label(
    ListNamaAbsensi_Grp,
    text="Sudah Absen:",
    bg="#FFFFFF",
    fg="#000000"
)
LabelSudahAbsen_Grp.grid(row=0, column=0, sticky='n')
##Label Aksi
LabelAksi_Lbl = tk.Label(
    AksiPilih_Grp,
    text="Aksi: ",
    fg="#000000",
    bg="#FFFFFF"
)
LabelAksi_Lbl.grid(row=0, column=0, sticky='n')
##List Nama yang Terdeteksi
listNamaTerdeteksi_Tk = tk.Listbox(
    ListNamaAbsensi_Grp,
    width=30,
    height=6,
    selectmode=tk.EXTENDED
    )
listNamaTerdeteksi_Tk.grid(row=1, column=0, sticky='n')
listNamaTerdeteksi_Tk.bind('<<ListboxSelect>>', AktifkanTombolHapus)
##Tombol Mulai Pengenalan
Pengenalan_BTN = tk.Button(TombolPengenalanDanStop_Grp,
                           width=15,
                           height=2,
                           text="Mulai",
                           command= AktifPengenalan
                           )
Pengenalan_BTN.grid(row=0, column=0)
##Tombol Stop Pengenalan
StopPengenalan_BTN = tk.Button(TombolPengenalanDanStop_Grp,
                           width=15,
                           height=2,
                           text="Stop",
                           command= NonaktifPengenalan
                           )
StopPengenalan_BTN.grid(row=0, column=1)
##Tambah Pengenalan
TambahPengenalan_BTN = tk.Button(TombolPengenalanDanStop_Grp,
                           width=15,
                           height=2,
                           text="Tambah",
                           command= TambahData
                           )
TambahPengenalan_BTN.grid(row=0, column=2)
##Simpan Absensi
SimpanAbsensi_BTN = tk.Button(TombolPengenalanDanStop_Grp,
                           width=15,
                           height=2,
                           text="Simpan",
                           command= SimpanAbsensi
                           )
SimpanAbsensi_BTN.grid(row=0, column=3)
##Tombol Aksi
AksiHapusNamaDariList_Grp = tk.Button(
    AksiPilih_Grp,
    text="Hapus",
    fg="#FF0000",
    padx=5,
    pady=10,
    state="disabled",
    command=HapusNamaDariList
)
AksiHapusNamaDariList_Grp.grid(row=1, column=0)

#Inisialisasi OpenCV Vide Capture
video_capture = cv2.VideoCapture(0)
video_capture.set(cv2.CAP_PROP_BUFFERSIZE, 2)


known_face_encodings = face_encoding
known_face_names = List_Nama


def capture():
    ret, frame = video_capture.read()

    rgb_small_frame = cv2.flip(frame, 1)
    cv2image = cv2.cvtColor(rgb_small_frame, cv2.COLOR_BGR2RGB) 


    global process_this_frame
    if process_this_frame and Mulai_Pengenalan:

        # Find all the faces and face encodings in the current frame of video
        face_locations = face_recognition.face_locations(cv2image)
        face_encodings = face_recognition.face_encodings(cv2image, face_locations)

        face_names = []
        for face_encoding in face_encodings:
            # See if the face is a match for the known face(s)
            matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
            name = "Unknown"


            face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
            best_match_index = np.argmin(face_distances)

            # Mendeteksi adanya Wajah Orang dari List
            if matches[best_match_index]:
                name = known_face_names[best_match_index]
                print("Halo " + name + ": " + "97%")
                #Ketika Nama Belum Ada pada List
                if name not in List_Nama_Terdeteksi:
                    # Menulis Nama Padea Variabel Memory
                    List_Nama_Terdeteksi.append(name)
                    # Menulis Nama Ke File
                    TulisNamaKeFile(name, LokasiFileNamaTerdeteksi)
                    print("Name Added")
                    listNamaTerdeteksi_Tk.insert("end", name)

            face_names.append(name)
    
    process_this_frame = not process_this_frame

    # Display the results
    img = Image.fromarray(cv2image)
    imgTk = ImageTk.PhotoImage(image=img)
    lmain.imgtk = imgTk
    lmain.configure(image=imgTk)
    lmain.after(10, capture)


# Runtime
if EncodePertama == True:
    LoadEncodingDataWajah()
    EncodePertama = not EncodePertama
capture()
window.mainloop()
video_capture.release()
cv2.destroyAllWindows()
